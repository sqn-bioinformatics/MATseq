"""Assemble and format the manuscript's tables from already-generated CSVs.

Earlier pipeline steps (MATseq.py -> src/model_training.py, src/pydeseq2.py,
src/go_term_analysis.py) write *raw* per-ligand / per-condition CSVs of
computed numbers. This module only opens those existing CSVs and reshapes them
into the final, correctly-named manuscript tables — no model fitting, no
DESeq2/GO calls, no new numbers.

Two groups of functions:

1. Main-text Table 2 — ``format_table2`` reads
   results/tables/table2_feature_set_benchmark.csv (written by STEP 6 of
   MATseq.py), long-form with columns condition, model, n_genes_mean,
   fit_seconds_{mean,std}, {accuracy,precision,recall,f1}_{mean,std}, and
   turns it into the manuscript table (csv + xlsx, best cell per metric bold).

2. Supplementary Tables S1-S12 — ``assemble_supplementary_table_{1..12}`` and
   the convenience wrapper ``assemble_supplementary_tables``. These read the
   raw DESeq2 / GO / feature-selection / nested-CV / validation CSVs (S5/S6
   read static wet-lab CSVs under data/supplementary_data/ instead) and emit
   the named Supplementary_Table_N files.

Raw inputs consumed (relative to the results directory, except S5/S6):
     S1: differential_gene_expression/<subset>/<ligand>_deseq2_results.csv
     S2: go_terms/<subset>/<ligand>_go_terms.csv   (per-ligand GO)
     S3: feature_selection/selected_vs_de_overlap_table.csv  (gene, in_de),
         ordered by feature-selection importance rank
     S4: go_terms/de_intersect_fs_go_terms.csv, go_terms/fs_only_go_terms.csv
     S5/S6: data/supplementary_data/Supplementary_Table_{5,6}.csv
     S7: feature_selection/mutual_information.csv (per-seed MI curves)
     S8: feature_selection/forest_kmeans.csv (ARI scan, best cell only)
     S9: tables/table2_feature_set_benchmark.csv (feature_selection condition)
     S10: nested_cv/supp_nested_cv_no_flapa.csv (feature_selection condition)
     S11/S12: validation/external_validation_{,no_flapa_}performance.csv
"""
from pathlib import Path

import pandas as pd
from sklearn.base import clone
from sklearn.model_selection import GridSearchCV, StratifiedKFold
from sklearn.pipeline import Pipeline as SkPipeline

from .config import CLASS_ORDER, FEATURE_SELECTION_CONFIG
from .feature_engineering import ColumnSelector, feature_pipeline, preprocessing_pipeline
from .model_training import make_score

# ----------------------------------------------------------------------------
# Table 2: feature-set ablation under leakage-free nested CV
# ----------------------------------------------------------------------------
def _condition_genes(trainer, condition, X_tr, y_tr, fold_seed, de_config,
                     subset, negative_control):
    """Derive the gene list for one condition on the OUTER-TRAINING fold only.

    Returns None for 'all_genes' (no column subsetting); otherwise a list of
    gene names frozen for this fold's inner tuning and outer-test scoring.
    """
    if condition == "all_genes":
        return None

    if condition == "feature_selection":
        fs = feature_pipeline(
            **FEATURE_SELECTION_CONFIG, random_state=fold_seed
        ).set_output(transform="pandas")
        fs.fit(X_tr, y_tr)
        return list(fs.get_feature_names_out())

    if condition == "random_selected":
        rng = np.random.default_rng(fold_seed)
        k = FEATURE_SELECTION_CONFIG["max_features"]
        cols = np.asarray(X_tr.columns)
        return sorted(rng.choice(cols, size=min(k, len(cols)), replace=False).tolist())

    if condition == "fs_plus_de":
        from .pydeseq2 import DESeq2  # local import: avoids any import cycle
        fs = feature_pipeline(
            **FEATURE_SELECTION_CONFIG, random_state=fold_seed
        ).set_output(transform="pandas")
        fs.fit(X_tr, y_tr)
        fs_genes = set(fs.get_feature_names_out())
        # DE genes from DESeq2 on the raw training counts of THIS fold only.
        y_tr_lab = pd.Series(trainer.decode_predictions(y_tr), index=X_tr.index)
        deseq = DESeq2(
            raw_counts=X_tr,
            sample_labels=y_tr_lab,
            padj_threshold=de_config["padj_threshold"],
            log2fc_threshold=de_config["log2fc_threshold"],
            n_cpus=de_config["n_cpus"],
            name=f"table2_fold{fold_seed}",
        )
        deseq.run_analysis(
            CLASS_ORDER[subset], negative_control=negative_control
        )
        de_genes = {str(g) for g in deseq.get_de_genes()}
        union = sorted(fs_genes | (de_genes & set(map(str, X_tr.columns))))
        return union

    raise ValueError(f"unknown condition: {condition}")

def _frozen_geneset_pipeline(model, gene_list):
    """preprocessing -> [ColumnSelector] -> clf, matching STEP 4b deployment
    order. XGBoost/RandomForest internal n_jobs forced to 1 so
    GridSearchCV(n_jobs=-1) owns all parallelism (avoids 48x48 thread
    oversubscription that stalled all_genes XGBoost)."""
    clf = clone(model)
    try:
        clf.set_params(n_jobs=1)
    except (ValueError, TypeError):
        pass
    steps = list(preprocessing_pipeline().steps)
    if gene_list is not None:
        steps.append(("select_genes", ColumnSelector(gene_list)))
    steps.append(("clf", clf))
    return SkPipeline(steps).set_output(transform="pandas")

def benchmark_feature_set_conditions(
    trainer,
    X,
    y,
    param_grids,
    output_dir,
    de_config,
    conditions=("all_genes", "feature_selection", "fs_plus_de", "random_selected"),
    outer_cv=5,
    inner_cv=3,
    scoring="f1_macro",
    subset="train_ligands",
    negative_control="negative_control",
):
    """Leakage-free nested-CV benchmark across feature-set conditions (Table 2).

    For each condition the gene set is derived on the outer-training fold only
    (see _condition_genes), frozen, then the inner 3-fold GridSearchCV tunes
    only classifier hyperparameters on that frozen set; the outer-test fold is
    scored once. Preprocessing (library-size norm, log1p, scaling) is refit in
    every fit. mean/std are taken across the outer folds (n=outer_cv).

    Writes long-form output_dir/table2_feature_set_benchmark.csv with columns:
    condition, model, n_genes_mean, and {metric}_{mean,std} for accuracy,
    precision, recall, f1. Returns that DataFrame.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    y_array = y.values if isinstance(y, pd.Series) else y
    trainer.label_encoder.fit(y_array)
    y_encoded = trainer.label_encoder.transform(y_array)

    min_class_count = np.bincount(y_encoded).min()
    if outer_cv > min_class_count:
        raise ValueError(
            f"outer_cv={outer_cv} exceeds smallest class count={min_class_count}"
        )

    outer = StratifiedKFold(
        n_splits=outer_cv, shuffle=True, random_state=trainer.random_state
    )
    fold_seeds = np.random.SeedSequence(trainer.random_state).generate_state(outer_cv)
    metric_keys = ["accuracy", "precision", "recall", "f1"]
    rows = []

    # Materialize the outer folds and their train/test slices once; reused
    # across every condition and model (the split is deterministic).
    folds = []
    for fold_idx, (train_idx, test_idx) in enumerate(
        outer.split(np.arange(len(X)), y_encoded)
    ):
        folds.append(
            (
                X.iloc[train_idx], X.iloc[test_idx],
                y_encoded[train_idx], y_encoded[test_idx],
            )
        )

    for cond in conditions:
        print(f"\n===== CONDITION: {cond} =====", flush=True)
        # Gene set is data-dependent, so derive once per outer fold and reuse
        # across all models within that fold.
        fold_genes = {}
        for fold_idx, (X_tr, _, y_tr, _) in enumerate(folds):
            genes = _condition_genes(
                    trainer,
                cond, X_tr, y_tr, int(fold_seeds[fold_idx]),
                de_config, subset, negative_control,
            )
            fold_genes[fold_idx] = genes
            print(f"  fold {fold_idx}: n_genes="
                  f"{'all' if genes is None else len(genes)}", flush=True)

        for model_name, model in trainer.models.items():
            per_fold = {k: [] for k in metric_keys}
            n_genes = []
            for fold_idx, (X_tr, X_te, y_tr, y_te) in enumerate(folds):
                genes = fold_genes[fold_idx]
                n_genes.append(X.shape[1] if genes is None else len(genes))
                pipe = _frozen_geneset_pipeline(model, genes)
                inner = StratifiedKFold(
                    n_splits=inner_cv, shuffle=True,
                    random_state=int(fold_seeds[fold_idx]),
                )
                gs = GridSearchCV(
                    pipe, param_grids[model_name], cv=inner,
                    scoring=scoring, n_jobs=-1, refit=True,
                )
                gs.fit(X_tr, y_tr, **trainer._fit_kwargs(model_name, y_tr))
                scores = make_score(y_te, gs.best_estimator_.predict(X_te))
                for k in metric_keys:
                    per_fold[k].append(scores[k])
            row = {"condition": cond, "model": model_name,
                   "n_genes_mean": float(np.mean(n_genes))}
            for k in metric_keys:
                row[f"{k}_mean"] = float(np.mean(per_fold[k]))
                row[f"{k}_std"] = float(np.std(per_fold[k], ddof=1))
            rows.append(row)
            print(f"  [{model_name}] f1={row['f1_mean']:.3f}±{row['f1_std']:.3f} "
                  f"acc={row['accuracy_mean']:.3f}±{row['accuracy_std']:.3f}",
                  flush=True)
            # persist incrementally so a wall-clock cutoff still yields data
            pd.DataFrame(rows).to_csv(
                output_dir / "table2_feature_set_benchmark.csv", index=False
            )

    df = pd.DataFrame(rows)
    df.to_csv(output_dir / "table2_feature_set_benchmark.csv", index=False)
    print(f"\nBenchmark saved: {output_dir / 'table2_feature_set_benchmark.csv'}",
          flush=True)
    return df

from .feature_engineering import best_forest_cell

# ----------------------------------------------------------------------------
# Supplementary table configuration
# ----------------------------------------------------------------------------
# Column order within each per-ligand DESeq2 block (DESeq2 native order).
DESEQ2_STATS = ["baseMean", "log2FoldChange", "lfcSE", "stat", "pvalue", "padj"]

# Display order of the ligand blocks across the wide S1 table. Each entry is
# (subset_dir, file_key, column_suffix): file_key is the stem in
# "<file_key>_deseq2_results.csv" inside differential_gene_expression/<subset_dir>/;
# column_suffix is what every stat column in that block is suffixed with, so the
# LPS block is explicitly labelled (previously it was left unsuffixed). LPS is
# taken from the train_ligands run; bacterial aliases map to HKEB/HKSA.
S1_LIGAND_ORDER = [
    ("train_ligands", "Fla-PA", "Fla-PA"),
    ("train_ligands", "LPS", "LPS"),
    ("train_ligands", "PGN", "PGN"),
    ("train_ligands", "R848", "R848"),
    ("train_ligands", "Pam3", "Pam3"),
    ("additional_ligands", "LTA", "LTA"),
    ("additional_ligands", "MPLA", "MPLA"),
    ("additional_ligands", "Pam2", "Pam2"),
    ("bacterial_ligands", "HK E.coli", "HKEB"),
    ("bacterial_ligands", "HK S.aureus", "HKSA"),
]

# Per-ligand GO files merged into S2 (the five main-ligand classes), given as
# (subset_dir, file_key, ligand_label).
S2_LIGAND_ORDER = [
    ("train_ligands", "Fla-PA", "Fla-PA"),
    ("train_ligands", "LPS", "LPS"),
    ("train_ligands", "PGN", "PGN"),
    ("train_ligands", "R848", "R848"),
    ("train_ligands", "Pam3", "Pam3"),
]

# Human-readable labels for the S4 by-condition GO table, mapped to the raw
# GO result file stem each is built from.
S4_CONDITIONS = [
    ("Intersect Feature Selection and Differentially Expressed Genes",
     "de_intersect_fs_go_terms.csv"),
    ("Feature Selection Genes not Differentially Expressed",
     "fs_only_go_terms.csv"),
]

# GO tables share this column layout (the leading index is dropped on write).
GO_COLUMNS = ["GO", "term", "class", "raw_pvalue", "fdr",
              "n_genes", "n_study", "ratio_in_study", "gene_symbols"]

# Display order and human-readable labels; label text is filled from the
# actual n_genes in the data so 450/958 never drift out of sync with config.
CONDITION_ORDER = ["all_genes", "feature_selection", "fs_plus_de", "random_selected"]
CONDITION_LABELS = {
    "all_genes": "All genes",
    "feature_selection": "Feature selection",
    "fs_plus_de": "Feature selection + DE",
    "random_selected": "Randomly selected",
}
MODEL_ORDER = ["LinearSVC", "SGDClassifier", "LogisticRegression",
               "RandomForest", "XGBoost"]
MODEL_DISPLAY_NAMES = {
    "LinearSVC": "Linear SVC",
    "SGDClassifier": "SGD",
    "RandomForest": "Random Forest",
}
# Column metrics in ML-paper order; (raw key, display header).
METRICS = [("accuracy", "Accuracy"), ("f1", "F1"),
           ("precision", "Precision"), ("recall", "Recall")]


def _cell(mean, std):
    return f"{mean:.3f} ± {std:.3f}"


def _condition_label(cond, n_genes):
    base = CONDITION_LABELS.get(cond, cond)
    if n_genes is None:
        return base
    return f"{base} ({int(round(n_genes)):,} genes)"


def format_table2(
    raw_csv,
    output_dir=None,
    bold_scope="column",
):
    """Reshape the raw benchmark CSV into the formatted Table 2.

    Args:
        raw_csv: path to table2_feature_set_benchmark.csv.
        output_dir: where to write outputs (defaults to the CSV's directory).
        bold_scope: 'column' bolds the single best mean per metric across the
            whole table; 'group' bolds the best per metric within each condition.

    Writes:
        table2_formatted.csv  — plain 'mean ± SD' cells, ** around bolded cells.
        table2.xlsx           — same content, real bold-cell formatting,
                                 friendly model names, a Training time column.
    Returns the formatted (long) DataFrame that backs both outputs.
    """
    raw_csv = Path(raw_csv)
    df = pd.read_csv(raw_csv)
    output_dir = Path(output_dir) if output_dir else raw_csv.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    df["condition"] = pd.Categorical(df["condition"], CONDITION_ORDER, ordered=True)
    present_models = [m for m in MODEL_ORDER if m in set(df["model"])]
    df["model"] = pd.Categorical(df["model"], present_models, ordered=True)
    df = df.sort_values(["condition", "model"]).reset_index(drop=True)

    # Which cell is best per metric (higher = better for all four metrics).
    best_mask = {}
    for key, _ in METRICS:
        col = f"{key}_mean"
        if bold_scope == "group":
            best_mask[key] = df.groupby("condition", observed=True)[col].transform("max")
        else:
            best_mask[key] = pd.Series(df[col].max(), index=df.index)

    rows = []
    for _, r in df.iterrows():
        row = {"Condition": _condition_label(r["condition"], r.get("n_genes_mean")),
               "Model": r["model"], "_fit_seconds_mean": r.get("fit_seconds_mean")}
        for key, header in METRICS:
            txt = _cell(r[f"{key}_mean"], r[f"{key}_std"])
            is_best = abs(r[f"{key}_mean"] - best_mask[key].iloc[r.name]) < 1e-9
            row[header] = f"**{txt}**" if is_best else txt
            row[f"_{key}_best"] = bool(is_best)
        rows.append(row)
    fmt = pd.DataFrame(rows)

    # Blank repeated condition labels so each group prints its label once.
    display = fmt.copy()
    display["Condition"] = display["Condition"].mask(
        display["Condition"].duplicated(), ""
    )
    cols = ["Condition", "Model"] + [h for _, h in METRICS]
    display[cols].to_csv(output_dir / "table2_formatted.csv", index=False)

    _write_xlsx(fmt, output_dir / "table2.xlsx")
    return fmt


def _write_xlsx(fmt, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    headers = [h for _, h in METRICS]
    wb = Workbook()
    ws = wb.active
    ws.append(["Condition", "Model"] + headers + ["Training time"])

    prev_cond = None
    for _, r in fmt.iterrows():
        cond = r["Condition"] if r["Condition"] != prev_cond else ""
        prev_cond = r["Condition"]
        model = MODEL_DISPLAY_NAMES.get(r["Model"], r["Model"])
        seconds = r["_fit_seconds_mean"]
        if pd.isna(seconds):
            training_time = ""
        elif seconds < 60:
            training_time = f"{seconds:.0f} s"
        elif seconds < 3600:
            training_time = f"{seconds / 60:.1f} min"
        else:
            training_time = f"{seconds / 3600:.1f} h"
        ws.append([cond, model] + [r[h].strip("*") for h in headers] + [training_time])
        for col_idx, (key, _) in enumerate(METRICS, start=3):
            if r[f"_{key}_best"]:
                ws.cell(row=ws.max_row, column=col_idx).font = Font(bold=True)

    wb.save(out_path)


# ============================================================================
# Supplementary tables S1-S4
# ============================================================================
def _read_go_table(path):
    """Read a raw GO CSV, dropping the leading unnamed index column."""
    df = pd.read_csv(path, index_col=0)
    return df.reset_index(drop=True)


def assemble_supplementary_table_1(results_dir, output_dir=None,
                                   filename="Supplementary_Table_1.csv"):
    """Assemble the wide per-ligand DESeq2 table (S1).

    Reads differential_gene_expression/<subset>/<ligand>_deseq2_results.csv for
    each ligand in S1_LIGAND_ORDER, suffixes every statistic column with the
    ligand label (so the LPS block is no longer unlabelled), and outer-joins the
    blocks on the gene identifier. The gene column is named ``gene``.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    de_dir = results_dir / "differential_gene_expression"
    merged = None
    for subset, file_key, suffix in S1_LIGAND_ORDER:
        fpath = de_dir / subset / f"{file_key}_deseq2_results.csv"
        if not fpath.exists():
            raise FileNotFoundError(f"S1: missing DESeq2 file {fpath}")
        block = pd.read_csv(fpath)
        # The raw file's first column is the gene id (header 'Geneid'); make it
        # a named 'gene' column and suffix the six statistic columns.
        block = block.rename(columns={block.columns[0]: "gene"})
        block = block.rename(columns={c: f"{c}_{suffix}" for c in DESEQ2_STATS
                                      if c in block.columns})
        merged = block if merged is None else merged.merge(block, on="gene", how="outer")

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    merged.to_csv(out_dir / filename, index=False)
    return merged


def assemble_supplementary_table_2(results_dir, output_dir=None,
                                   filename="Supplementary_Table_2.csv"):
    """Assemble the per-ligand GO enrichment table (S2).

    Concatenates the five main-ligand GO tables (S2_LIGAND_ORDER) with a leading
    ``Ligand`` column, sorted by FDR. No stray index column is written.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    go_dir = results_dir / "go_terms"
    frames = []
    for subset, file_key, label in S2_LIGAND_ORDER:
        fpath = go_dir / subset / f"{file_key}_go_terms.csv"
        if not fpath.exists():
            raise FileNotFoundError(f"S2: missing GO file {fpath}")
        df = _read_go_table(fpath)
        df.insert(0, "Ligand", label)
        frames.append(df)
    merged = pd.concat(frames, axis=0, ignore_index=True)
    merged = merged.sort_values("fdr", ascending=True).reset_index(drop=True)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    merged.to_csv(out_dir / filename, index=False)
    return merged


def assemble_supplementary_table_3(results_dir, output_dir=None,
                                   filename="Supplementary_Table_3.csv",
                                   include_rank=True):
    """Assemble the selected-genes table (S3).

    Reads feature_selection/selected_vs_de_overlap_table.csv (columns
    ``gene``, ``in_de`` and, from the current pipeline, ``rank`` and
    ``importance``) and emits ``Gene`` and ``Differentially_Expressed``.
    When the source carries a ``rank`` column the rows are ordered by
    ExtraTrees feature importance (rank 1 = most important) and the optional
    ``Rank`` column reflects that true importance rank; otherwise Rank falls
    back to row order. Set include_rank=False to omit the Rank column.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "feature_selection" / "selected_vs_de_overlap_table.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S3: missing feature-selection file {fpath}")
    src = pd.read_csv(fpath)
    if "rank" in src.columns:
        src = src.sort_values("rank").reset_index(drop=True)
    out = pd.DataFrame({
        "Gene": src["gene"],
        "Differentially_Expressed": src["in_de"].astype(bool),
    })
    if include_rank:
        if "rank" in src.columns:
            out.insert(0, "Rank", src["rank"].astype(int).values)
        else:
            out.insert(0, "Rank", range(1, len(out) + 1))

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_4(results_dir, output_dir=None,
                                   filename="Supplementary_Table_4.csv"):
    """Assemble the by-condition GO enrichment table (S4).

    Concatenates the two gene-set GO tables (S4_CONDITIONS) with a leading
    ``Condition`` column. No stray index column is written.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    go_dir = results_dir / "go_terms"
    frames = []
    for label, fname in S4_CONDITIONS:
        fpath = go_dir / fname
        if not fpath.exists():
            raise FileNotFoundError(f"S4: missing GO file {fpath}")
        df = _read_go_table(fpath)
        df.insert(0, "Condition", label)
        frames.append(df)
    merged = pd.concat(frames, axis=0, ignore_index=True)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    merged.to_csv(out_dir / filename, index=False)
    return merged


def assemble_supplementary_table_5(results_dir, output_dir=None,
                                   filename="Supplementary_Table_5.csv"):
    """Assemble the LPS OD630nm titration table (S5) from the wet-lab source."""
    fpath = Path(__file__).resolve().parent.parent / "data" / "supplementary_data" / "Supplementary_Table_5.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S5: missing wet-lab source file {fpath}")
    out = pd.read_csv(fpath).rename(
        columns={"Concentration_(EU_mL)": "Concentration_LPS_(eq_mL)"}
    )

    out_dir = Path(output_dir) if output_dir else Path(results_dir) / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_6(results_dir, output_dir=None,
                                   filename="Supplementary_Table_6.csv"):
    """Assemble the Pam3 OD630nm titration table (S6) from the wet-lab source."""
    fpath = Path(__file__).resolve().parent.parent / "data" / "supplementary_data" / "Supplementary_Table_6.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S6: missing wet-lab source file {fpath}")
    out = pd.read_csv(fpath).rename(
        columns={"Concentration_(ng_mL)": "Concentration_Pam3_(ng_mL)"}
    )

    out_dir = Path(output_dir) if output_dir else Path(results_dir) / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_7(results_dir, output_dir=None,
                                   filename="Supplementary_Table_7.csv"):
    """Assemble the per-seed mutual-information table (S7).

    Reads feature_selection/mutual_information.csv (rank, mi_sorted,
    mi_sorted_seed_<seed>...) and renames to gene_rank, MI_mean,
    MI_seed_<seed>.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "feature_selection" / "mutual_information.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S7: missing mutual information file {fpath}")
    src = pd.read_csv(fpath)
    seed_cols = [c for c in src.columns if c.startswith("mi_sorted_seed_")]
    seed_names = [c.removeprefix("mi_sorted_seed_") for c in seed_cols]
    rename = {"rank": "gene_rank", "mi_sorted": "MI_mean"}
    rename.update({c: f"MI_seed_{s}" for c, s in zip(seed_cols, seed_names)})
    out = src.rename(columns=rename)
    out = out[["gene_rank"] + [f"MI_seed_{s}" for s in seed_names] + ["MI_mean"]]

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_8(results_dir, output_dir=None,
                                   filename="Supplementary_Table_8.csv"):
    """Assemble the forest/k-means ARI-vs-gene-count table (S8).

    Reads feature_selection/forest_kmeans.csv, filters to the winning
    (n_estimators, max_depth) cell (feature_engineering.best_forest_cell), and
    renames n_selected/ari_mean/ari_std/ari_seed_<seed> to n_selected_genes/
    ARI_mean/ARI_std/ARI_seed_<seed>.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "feature_selection" / "forest_kmeans.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S8: missing forest/k-means scan file {fpath}")
    scan = pd.read_csv(fpath)
    cell = best_forest_cell(scan)
    seed_cols = [c for c in cell.columns if c.startswith("ari_seed_")]
    seed_names = [c.removeprefix("ari_seed_") for c in seed_cols]
    rename = {"n_selected": "n_selected_genes", "ari_mean": "ARI_mean",
              "ari_std": "ARI_std"}
    rename.update({c: f"ARI_seed_{s}" for c, s in zip(seed_cols, seed_names)})
    out = cell.rename(columns=rename)
    cols = (["n_selected_genes"] + [f"ARI_seed_{s}" for s in seed_names]
            + ["ARI_mean", "ARI_std"])
    out = out[cols].reset_index(drop=True)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


# Narrow nested-CV column set shared by S9 and S10 (drops the per-fold
# precision/recall columns and every pooled_* metric but f1).
NESTED_CV_COLUMNS = [
    "model", "accuracy_mean", "accuracy_std",
    "balanced_accuracy_mean", "balanced_accuracy_std",
    "f1_mean", "f1_std", "f1_weighted_mean", "f1_weighted_std", "pooled_f1",
]


def assemble_supplementary_table_9(results_dir, output_dir=None,
                                   filename="Supplementary_Table_9.csv"):
    """Assemble the with-Fla-PA nested-CV summary table (S9).

    Reads tables/table2_feature_set_benchmark.csv (all 4 gene-set conditions),
    keeps only the feature_selection condition rows and the narrow S9 column
    set.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "tables" / "table2_feature_set_benchmark.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S9: missing benchmark file {fpath}")
    src = pd.read_csv(fpath)
    src = src[src["condition"] == "feature_selection"]
    present_models = [m for m in MODEL_ORDER if m in set(src["model"])]
    src["model"] = pd.Categorical(src["model"], present_models, ordered=True)
    out = src.sort_values("model")[NESTED_CV_COLUMNS].reset_index(drop=True)
    out["model"] = out["model"].astype(str)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_10(results_dir, output_dir=None,
                                    filename="Supplementary_Table_10.csv"):
    """Assemble the no-Fla-PA nested-CV summary table (S10).

    Reads nested_cv/supp_nested_cv_no_flapa.csv (all 4 gene-set conditions),
    keeps only the feature_selection condition rows and the narrow S10 column
    set.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "nested_cv" / "supp_nested_cv_no_flapa.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S10: missing benchmark file {fpath}")
    src = pd.read_csv(fpath)
    src = src[src["condition"] == "feature_selection"]
    present_models = [m for m in MODEL_ORDER if m in set(src["model"])]
    src["model"] = pd.Categorical(src["model"], present_models, ordered=True)
    out = src.sort_values("model")[NESTED_CV_COLUMNS].reset_index(drop=True)
    out["model"] = out["model"].astype(str)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_11(results_dir, output_dir=None,
                                    filename="Supplementary_Table_11.csv"):
    """Assemble the external validation performance table (S11).

    Passes validation/external_validation_performance.csv through unchanged
    (columns already match: gene_set, model, accuracy, balanced_accuracy,
    precision, recall, f1, f1_weighted).

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "validation" / "external_validation_performance.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S11: missing external validation file {fpath}")
    out = pd.read_csv(fpath)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_table_12(results_dir, output_dir=None,
                                    filename="Supplementary_Table_12.csv"):
    """Assemble the no-Fla-PA external validation performance table (S12).

    Passes validation/external_validation_no_flapa_performance.csv through
    unchanged.

    Returns the assembled DataFrame.
    """
    results_dir = Path(results_dir)
    fpath = results_dir / "validation" / "external_validation_no_flapa_performance.csv"
    if not fpath.exists():
        raise FileNotFoundError(f"S12: missing external validation file {fpath}")
    out = pd.read_csv(fpath)

    out_dir = Path(output_dir) if output_dir else results_dir / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.to_csv(out_dir / filename, index=False)
    return out


def assemble_supplementary_tables(results_dir, output_dir=None):
    """Regenerate Supplementary Tables S1-S12 from raw pipeline outputs.

    Returns a dict mapping table name to the assembled DataFrame.
    """
    return {
        "S1": assemble_supplementary_table_1(results_dir, output_dir),
        "S2": assemble_supplementary_table_2(results_dir, output_dir),
        "S3": assemble_supplementary_table_3(results_dir, output_dir),
        "S4": assemble_supplementary_table_4(results_dir, output_dir),
        "S5": assemble_supplementary_table_5(results_dir, output_dir),
        "S6": assemble_supplementary_table_6(results_dir, output_dir),
        "S7": assemble_supplementary_table_7(results_dir, output_dir),
        "S8": assemble_supplementary_table_8(results_dir, output_dir),
        "S9": assemble_supplementary_table_9(results_dir, output_dir),
        "S10": assemble_supplementary_table_10(results_dir, output_dir),
        "S11": assemble_supplementary_table_11(results_dir, output_dir),
        "S12": assemble_supplementary_table_12(results_dir, output_dir),
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "supp":
        results_dir = sys.argv[2] if len(sys.argv) > 2 else "results"
        tables = assemble_supplementary_tables(results_dir)
        for name, df in tables.items():
            print(f"{name}: {df.shape}  cols={list(df.columns)[:6]}")
    else:
        raw = sys.argv[1] if len(sys.argv) > 1 else "results/tables/table2_feature_set_benchmark.csv"
        out = format_table2(raw)
        print(out[["Condition", "Model"] + [h for _, h in METRICS]].to_string(index=False))
